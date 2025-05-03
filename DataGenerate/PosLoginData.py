import openpyxl

def dataGenerator():
    wk = openpyxl.load_workbook("C:/Users/rahul/Downloads/Login.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    data_list = []

    for i in range(1, r+1):
        mobile_no = sh.cell(i, 1).value
        otp = sh.cell(i, 2).value
        data_list.append((mobile_no, otp))

    print(data_list)  # Optional: Print the data list for verification
    return data_list